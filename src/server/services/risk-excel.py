#!/usr/bin/env python3
"""
Magneto — Risk Assessment Excel Generator (Python/openpyxl)
Called by risk-excel.ts via child_process.execFile
Input: JSON via stdin  { devices: [...], label: "..." }
Output: xlsx bytes via stdout
"""
import sys, json, io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Paleta exata dos arquivos de referência ───────────────────────────────────
C = dict(
    PURPLE_HARD = "461E5F",   # Header hardening seção
    BLUE_INFRA  = "1F4E79",   # Header infraestrutura seção  
    RED         = "FF0000",   # ALTO
    AMBER       = "FFC000",   # MÉDIA
    GREEN       = "70AD47",   # BAIXA
    GRAY        = "D9D9D9",   # N/A fill
    WHITE       = "FFFFFF",
    BLACK       = "000000",
    GRAY_TEXT   = "555555",
    # Assessment colors
    ASSESS_PURPLE = "6B0AC9", # fill headers assessment
    ASSESS_TITLE  = "2E0060", # título assess (font)
    ASSESS_DARK1  = "D9D9D9", # row alt 1
    ASSESS_DARK2  = "BFBFBF", # row alt 2
)

def sf(rgb):
    return PatternFill("solid", fgColor=rgb)

def fnt(rgb=None, bold=False, sz=10, name="Calibri"):
    return Font(name=name, color=rgb or C["BLACK"], bold=bold, size=sz)

def alh(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

_thin = Side(style="thin", color="000000")
_hair = Side(style="hair", color="AAAAAA")

def bdr_thin():
    return Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

def bdr_hair():
    return Border(left=_hair, right=_hair, top=_hair, bottom=_hair)

# ── MATRIZ DE RISCOS — uma sheet por device ───────────────────────────────────
def build_matriz_sheet(wb, dev, idx):
    hostname = dev.get("hostname") or f"Device_{idx+1}"
    name = hostname[:31]
    # Evitar nomes duplicados
    existing = [ws.title for ws in wb.worksheets]
    if name in existing:
        name = name[:28] + f"_{idx}"
    ws = wb.create_sheet(name)

    # Larguras exatas do arquivo de referência
    ws.column_dimensions['A'].width = 14.43
    ws.column_dimensions['B'].width = 41.71
    ws.column_dimensions['C'].width = 21.86
    ws.column_dimensions['D'].width = 14.43
    ws.column_dimensions['E'].width = 55.86

    # ── Legenda (rows 3-7) ────────────────────────────────────────────────────
    ws.merge_cells('B3:E3')
    c = ws['B3']
    c.value = 'LEGENDA'; c.fill = sf(C["PURPLE_HARD"])
    c.font = fnt(C["WHITE"], sz=11); c.alignment = alh("center")

    for row, val, fill_rgb, fc in [
        (4, '✕  ALTO',  C["RED"],   C["WHITE"]),
        (5, '⚠  MÉDIA', C["AMBER"], C["BLACK"]),
        (6, '✔  BAIXA', C["GREEN"], C["WHITE"]),
        (7, 'N/A',      C["GRAY"],  C["GRAY_TEXT"]),
    ]:
        ws.merge_cells(f'B{row}:E{row}')
        c = ws.cell(row, 2)
        c.value = val; c.fill = sf(fill_rgb)
        c.font = fnt(fc, sz=11); c.alignment = alh("center")

    # ── Nome do device + título hardening (rows 9-10) ─────────────────────────
    ws.merge_cells('B9:E9')
    c = ws['B9']
    c.value = hostname
    c.fill = sf(C["PURPLE_HARD"]); c.font = fnt(C["WHITE"], sz=11)
    c.alignment = alh("center")
    ws.row_dimensions[9].height = 16

    ws.merge_cells('B10:E10')
    c = ws['B10']
    c.value = 'HARDENING: Requisitos de Segurança'
    c.fill = sf(C["PURPLE_HARD"]); c.font = fnt(C["WHITE"], sz=11)
    c.alignment = alh("center")
    ws.row_dimensions[10].height = 16

    # ── Renderizar items ──────────────────────────────────────────────────────
    cur_row = 11
    items = dev.get("items", [])
    
    # Estado da seção atual
    cur_section_color = C["PURPLE_HARD"]  # alterna entre PURPLE e BLUE
    section_count = 0
    in_hardening = True  # começa no bloco de hardening

    for item in items:
        status = item.get("status", "")
        
        if status == "SECTION":
            section_name = item.get("item", "")
            # Detectar se mudou para infra (contém "INFRAESTRUTURA")
            if "INFRAESTRUTURA" in section_name.upper() or "ROTEAMENTO" in section_name.upper():
                in_hardening = False
            
            color = C["BLUE_INFRA"] if not in_hardening else C["PURPLE_HARD"]
            
            # Título da seção com colunas
            for col in [2, 3, 4, 5]:
                c = ws.cell(cur_row, col)
                c.fill = sf(color)
                c.font = fnt(C["WHITE"], sz=11)
                c.alignment = alh()
            
            ws.cell(cur_row, 2).value = section_name
            
            # Se tem sub-colunas (STATUS, RISCO, OBSERVAÇÃO)
            section_name_upper = section_name.upper()
            has_cols = any(k in section_name_upper for k in ['AUTENTICAÇÃO', 'CRIPTOGRAFIA', 'ACESSO', 'GERÊNCIA', 
                                                               'ROTEAMENTO', 'GATEWAY', 'SERVIÇOS', 'SWITCHING', 
                                                               'PORT-CHANNEL', 'INFRAESTRUTURA', 'HARDENING'])
            if has_cols:
                ws.cell(cur_row, 3).value = 'STATUS'
                ws.cell(cur_row, 4).value = 'RISCO'
                ws.cell(cur_row, 5).value = 'OBSERVAÇÃO'
            
            ws.row_dimensions[cur_row].height = 15
            cur_row += 1
            section_count += 1
            continue

        # ── Item normal ───────────────────────────────────────────────────────
        risco = item.get("risco", "")
        obs   = item.get("obs", "")

        # Determinar fill do campo RISCO
        if '✕' in str(risco) or 'ALTO' in str(risco).upper():
            risco_fill = sf(C["RED"]); risco_fc = C["BLACK"]
        elif '⚠' in str(risco) or 'MÉDIA' in str(risco).upper() or 'MEDIO' in str(risco).upper():
            risco_fill = sf(C["AMBER"]); risco_fc = C["BLACK"]
        elif '✔' in str(risco) or 'BAIXA' in str(risco).upper() or 'BAIXO' in str(risco).upper():
            risco_fill = sf(C["GREEN"]); risco_fc = C["WHITE"]
        else:  # N/A
            risco_fill = sf(C["GRAY"]); risco_fc = C["BLACK"]

        # Determinar fill do STATUS
        stat_str = str(status).upper()
        if stat_str == 'SIM':
            stat_val = 'SIM'
        elif stat_str == 'NÃO' or stat_str == 'NAO':
            stat_val = 'NÃO'
        elif stat_str == 'PARCIAL':
            stat_val = 'PARCIAL'
        else:
            stat_val = status

        # Item (col B)
        cb = ws.cell(cur_row, 2)
        cb.value = item.get("item", "")
        cb.font = fnt(C["BLACK"], sz=10); cb.alignment = alh("left")
        cb.border = bdr_thin()

        # STATUS (col C)
        cc = ws.cell(cur_row, 3)
        cc.value = stat_val
        cc.font = fnt(C["BLACK"], bold=True, sz=10)
        cc.alignment = alh("center")
        cc.border = bdr_thin()

        # RISCO (col D)
        cd = ws.cell(cur_row, 4)
        # Formatar risco como no referência
        if '✕' in str(risco):   cd.value = '✕  ALTO'
        elif '⚠' in str(risco): cd.value = '⚠  MÉDIA'
        elif '✔' in str(risco): cd.value = '✔  BAIXA'
        elif risco:              cd.value = risco
        else:                    cd.value = 'N/A'
        cd.fill = risco_fill
        cd.font = fnt(risco_fc, bold=True, sz=10)
        cd.alignment = alh("center")
        cd.border = bdr_thin()

        # OBSERVAÇÃO (col E)
        ce = ws.cell(cur_row, 5)
        ce.value = obs
        ce.font = fnt(C["BLACK"], sz=10); ce.alignment = alh("left", wrap=True)
        ce.border = bdr_thin()

        ws.row_dimensions[cur_row].height = 15
        cur_row += 1

    return ws

# ── ASSESSMENT — structure from Assessment.xlsx ────────────────────────────────
def build_assessment_wb(devices, label):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    NC = "Não configurado"
    PFILL = sf(C["ASSESS_PURPLE"])
    GLFIL = sf(C["ASSESS_DARK1"])
    GMFIL = sf(C["ASSESS_DARK2"])
    TITLE_FONT = Font(name="Calibri", bold=True, size=16, color=C["ASSESS_TITLE"])
    HDR_FONT   = Font(name="Calibri", bold=True, size=10, color=C["WHITE"])
    DAT_FONT   = Font(name="Calibri", size=9)
    _bdr = bdr_thin()

    def write_header(ws, title, sub, nc):
        ws.row_dimensions[1].height = 26
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=nc+1)
        c = ws.cell(1, 2)
        c.value = title; c.font = TITLE_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        
        ws.row_dimensions[2].height = 18
        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=nc+1)
        c = ws.cell(2, 2)
        c.value = sub; c.fill = PFILL; c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        for col in range(3, nc+2):
            ws.cell(2, col).fill = PFILL

    def write_col_headers(ws, row, headers):
        ws.row_dimensions[row].height = 16
        for i, h in enumerate(headers):
            c = ws.cell(row, i+2)
            c.value = h; c.fill = PFILL; c.border = _bdr
            c.font = HDR_FONT
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def write_hostname_block(ws, row, hostname, nc):
        ws.row_dimensions[row].height = 16
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=nc+1)
        c = ws.cell(row, 2)
        c.value = hostname; c.fill = PFILL; c.border = _bdr
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")

    def write_data_rows(ws, start_row, rows, nc):
        for ri, row_data in enumerate(rows):
            r = start_row + ri
            ws.row_dimensions[r].height = 14
            fill = GLFIL if ri % 2 == 0 else GMFIL
            for ci, val in enumerate(row_data):
                c = ws.cell(r, ci+2)
                v = NC if (val is None or val == "") else str(val)
                c.value = v; c.fill = fill; c.border = _bdr
                c.font = DAT_FONT
                c.alignment = Alignment(vertical="center")
        return start_row + len(rows)

    def add_sheet(name, title, headers, rows, widths):
        ws = wb.create_sheet(name)
        nc = len(headers)
        ws.column_dimensions['A'].width = 2
        for i, w in enumerate(widths):
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(i+2)].width = w
        write_header(ws, title, label, nc)
        write_col_headers(ws, 3, headers)
        if rows:
            write_data_rows(ws, 4, rows, nc)
        return ws

    def add_sheet_by_host(name, title, headers, device_rows, widths):
        ws = wb.create_sheet(name)
        nc = len(headers)
        ws.column_dimensions['A'].width = 2
        from openpyxl.utils import get_column_letter
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(i+2)].width = w
        write_header(ws, title, label, nc)
        row = 3
        for block in device_rows:
            hostname = block.get("hostname", "")
            rows = block.get("rows", [])
            write_hostname_block(ws, row, hostname, nc)
            row += 1
            write_col_headers(ws, row, headers)
            row += 1
            data = rows if rows else [[NC]*nc]
            row = write_data_rows(ws, row, data, nc) + 1
        return ws

    def v(d, key):
        val = d.get(key)
        return NC if (val is None or val == "") else str(val)

    def vl(d, key):
        val = d.get(key, [])
        return val if isinstance(val, list) else []

    # Helper to get vendor label
    VENDOR_LABEL = {
        "cisco_ios":   "Cisco", "cisco_nxos": "Cisco",
        "dell_os10":   "Dell",  "hpe_comware": "HP", "huawei_vrp": "Huawei",
    }
    OS_TYPE_LABEL = {
        "cisco_ios":   "IOS",    "cisco_nxos": "NX-OS",
        "dell_os10":   "OS10",   "hpe_comware": "HP Comware", "huawei_vrp": "Huawei VRP",
    }

    # ── Inventário ────────────────────────────────────────────────────────────
    inv_rows = []
    for d in devices:
        fab = VENDOR_LABEL.get(d.get("vendor",""), "Cisco")
        inv_rows.append([v(d,"hostname"), "Switch", fab, v(d,"model"), v(d,"serial") or NC])
    add_sheet("Inventário","Inventário",
              ["Hostname","Tipo","Fabricante","Part Number","Serial Number"],
              inv_rows, [22,14,14,22,22])

    # ── EOL ───────────────────────────────────────────────────────────────────
    eol_rows = []
    for d in devices:
        link = "https://www.cisco.com/c/en/us/products/switches/"
        eol_rows.append([v(d,"hostname"), v(d,"model"), NC, link])
    add_sheet("EOL","End of Life",
              ["Hostname","Part Number","EOL","Observação / Link"],
              eol_rows, [22,22,14,60])

    # ── Versões de Softwares ──────────────────────────────────────────────────
    ver_rows = []
    for d in devices:
        fab = VENDOR_LABEL.get(d.get("vendor",""), "Cisco")
        tipo = OS_TYPE_LABEL.get(d.get("vendor",""), "IOS")
        ver_rows.append([v(d,"hostname"), fab, v(d,"model"), tipo, v(d,"osVersion")])
    add_sheet("Versões de Softwares","Versões de Software",
              ["Hostname","Fabricante","Modelo","Tipo Imagem","Versão"],
              ver_rows, [22,14,22,14,20])

    # ── Software Recomendados ─────────────────────────────────────────────────
    rec_rows = []
    for d in devices:
        fab = VENDOR_LABEL.get(d.get("vendor",""), "Cisco")
        tipo = OS_TYPE_LABEL.get(d.get("vendor",""), "IOS")
        rec_rows.append([v(d,"hostname"), fab, v(d,"model"), tipo, v(d,"osVersion"), NC])
    add_sheet("Software Recomendados","Versões de Software Recomendadas",
              ["Hostname","Fabricante","Modelo","Tipo Imagem","Versão atual","Versão recomendada"],
              rec_rows, [22,14,22,14,20,22])

    # ── IP DE GERÊNCIA ────────────────────────────────────────────────────────
    ws_ip = wb.create_sheet("IP DE GERÊNCIA")
    nc = 5
    ws_ip.column_dimensions['A'].width = 2
    from openpyxl.utils import get_column_letter
    for i, w in enumerate([22,18,16,18,18]):
        ws_ip.column_dimensions[get_column_letter(i+2)].width = w
    write_header(ws_ip, "IP de Gerência", label, nc)
    row = 3
    # OOB block
    ws_ip.merge_cells(start_row=row, start_column=2, end_row=row, end_column=nc+1)
    c = ws_ip.cell(row, 2)
    c.value = "OUT-OF-BAND (mgmt0)"; c.fill = sf("2E4057"); c.border = _bdr
    c.font = Font(name="Calibri", bold=True, size=11, color=C["WHITE"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    row += 1
    write_col_headers(ws_ip, row, ["Hostname","IP","Mask","Gateway","Interface"])
    row += 1
    oob_rows = [[v(d,"hostname"), v(d,"ip"), NC, NC, v(d,"mgmtIntf") or NC] for d in devices]
    if oob_rows:
        row = write_data_rows(ws_ip, row, oob_rows, nc) + 1
    else:
        write_data_rows(ws_ip, row, [[NC,NC,NC,NC,NC]], nc); row += 2
    # IN-BAND block
    ws_ip.merge_cells(start_row=row, start_column=2, end_row=row, end_column=nc+1)
    c = ws_ip.cell(row, 2)
    c.value = "IN-BAND (Vlan / Interface)"; c.fill = sf("1B4332"); c.border = _bdr
    c.font = Font(name="Calibri", bold=True, size=11, color=C["WHITE"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    row += 1
    write_col_headers(ws_ip, row, ["Hostname","IP","Mask","Gateway","Interface"])
    row += 1
    ib_rows = [[v(d,"hostname"), v(d,"ip"), NC, NC, NC] for d in devices]
    write_data_rows(ws_ip, row, ib_rows if ib_rows else [[NC,NC,NC,NC,NC]], nc)

    # ── CDP ───────────────────────────────────────────────────────────────────
    cdp_data = [{"hostname":v(d,"hostname"),"rows":[[r.get("devId",NC),r.get("ip",NC),r.get("localIf",NC),r.get("hold",NC),r.get("cap",NC),r.get("plat",NC),r.get("remIf",NC)] for r in vl(d,"cdp")] if vl(d,"cdp") else []} for d in devices]
    add_sheet_by_host("CDP","CDP - Cisco Discovery Protocol",
                      ["Device ID","IP","Local Intf","Hold-time","Capability","Platform","Port ID"],
                      cdp_data, [40,16,16,10,20,22,16])

    # ── LLDP ──────────────────────────────────────────────────────────────────
    lldp_data = [{"hostname":v(d,"hostname"),"rows":[[r.get("devId",NC),r.get("localIf",NC),r.get("hold",NC),r.get("cap",NC),r.get("remIf",NC)] for r in vl(d,"lldp")] if vl(d,"lldp") else []} for d in devices]
    add_sheet_by_host("LLDP","LLDP - Link Layer Discovery Protocol",
                      ["Device ID","Local Intf","Hold-time","Capability","Port ID"],
                      lldp_data, [40,16,10,16,16])

    # ── VLANs ─────────────────────────────────────────────────────────────────
    vlan_data = [{"hostname":v(d,"hostname"),"rows":[[r.get("id",NC),r.get("name",NC),r.get("status",NC)] for r in vl(d,"vlans")]} for d in devices]
    add_sheet_by_host("VLANS","VLANs",["VLAN ID","Name","Status"],vlan_data,[10,30,12])

    # ── VTP ───────────────────────────────────────────────────────────────────
    vtp_data = [{"hostname":v(d,"hostname"),"rows":[[v(d,"vtpVer"),v(d,"vtpDomain"),v(d,"vtpMode"),v(d,"vtpVlans"),v(d,"vtpRev"),v(d,"vtpPwd")]]} for d in devices]
    add_sheet_by_host("VTP","VLAN Trunking Protocol (VTP)",["VTP Versão","VTP Domain Name","VTP Mode","Nº VLANs","Config Revision","Password"],vtp_data,[10,22,14,10,14,22])

    # ── STP ───────────────────────────────────────────────────────────────────
    stp_data = [{"hostname":v(d,"hostname"),"rows":[[r.get("vlan",NC),str(r.get("rootPri",NC))+" "+str(r.get("rootMac",NC)),r.get("cost",NC),r.get("rootPort",NC)] for r in vl(d,"stp")]} for d in devices]
    add_sheet_by_host("STP","Spanning Tree Protocol",["VLAN","Root Bridge ID","Cost","Root Port"],stp_data,[14,28,10,14])

    # ── INT_VLAN ──────────────────────────────────────────────────────────────
    ivlan_data = [{"hostname":v(d,"hostname"),"rows":[[r.get("vid",NC),r.get("desc",NC),r.get("ip",NC),r.get("mask",NC),", ".join(r.get("helper",[])) or NC] for r in vl(d,"intVlan")]} for d in devices]
    add_sheet_by_host("INT_VLAN","Interfaces VLAN (SVIs)",["Vlan ID","Description","IP","Mask","IP Helper"],ivlan_data,[12,34,16,16,24])

    # ── HSRP ──────────────────────────────────────────────────────────────────
    hsrp_data = [{"hostname":v(d,"hostname"),"rows":[[r.get("platform",NC),r.get("intf",NC),r.get("grp",NC),r.get("pri",NC),r.get("p",NC),r.get("state",NC),r.get("active",NC),r.get("standby",NC),r.get("vip",NC)] for r in vl(d,"hsrp")]} for d in devices]
    add_sheet_by_host("HSRP","HSRP - Hot Standby Router Protocol",["Platform","Interface","Grp","Pri","P","State","Active","Standby","Virtual IP"],hsrp_data,[10,14,8,8,6,10,16,16,16])

    # ── VRRP ──────────────────────────────────────────────────────────────────
    vrrp_data = [{"hostname":v(d,"hostname"),"rows":[[r.get("vrid",NC),r.get("intf",NC),r.get("state",NC),r.get("type",NC),r.get("vip",NC)] for r in vl(d,"vrrp")]} for d in devices]
    add_sheet_by_host("VRRP","VRRP - Virtual Router Redundancy Protocol",["VRID","Interface","State","Type","Virtual IP"],vrrp_data,[10,16,12,12,16])

    # ── GLBP ──────────────────────────────────────────────────────────────────
    glbp_data = [{"hostname":v(d,"hostname"),"rows":[[r.get("intf",NC),r.get("grp",NC),r.get("pri",NC),r.get("state",NC),r.get("vip",NC),r.get("active",NC),r.get("standby",NC)] for r in vl(d,"glbp")]} for d in devices]
    add_sheet_by_host("GLBP","GLBP - Gateway Load Balancing Protocol",["Interface","Grp","Pri","State","Virtual IP","Active Router","Standby Router"],glbp_data,[14,8,8,10,16,16,16])

    # ── INT_STATUS ────────────────────────────────────────────────────────────
    ist_data = [{"hostname":v(d,"hostname"),"rows":[[r.get("port",NC),r.get("desc",NC),r.get("status",NC),r.get("vlan",NC),r.get("duplex",NC),r.get("speed",NC),r.get("type",NC)] for r in vl(d,"intSt")]} for d in devices]
    add_sheet_by_host("INT_STATUS","Status das Interfaces",["Port","Description","Status","Vlan","Duplex","Speed","Type"],ist_data,[12,35,12,10,8,8,20])

    # ── PORT-CHANNEL ──────────────────────────────────────────────────────────
    pc_data = [{"hostname":v(d,"hostname"),"rows":[[r.get("po",NC),r.get("members",NC),r.get("vizinho",NC),NC,r.get("portasRemotas",NC),r.get("status",NC),r.get("proto",NC)] for r in vl(d,"portch")]} for d in devices]
    add_sheet_by_host("PORT-CHANNEL","Port-Channel / EtherChannel",["Port-Channel Local","Portas Local","Vizinho","Port-Channel Remoto","Portas Remotas","Status","Protocol"],pc_data,[18,40,36,20,20,10,10])

    # ── TRUNK ─────────────────────────────────────────────────────────────────
    trunk_data = [{"hostname":v(d,"hostname"),"rows":[[r.get("port",NC),r.get("mode",NC),r.get("encap",NC),r.get("status",NC),r.get("vlans",NC),r.get("native",NC)] for r in vl(d,"trunk")]} for d in devices]
    add_sheet_by_host("TRUNK","Interfaces Trunk",["PORT","MODE","ENCAPSULATION","STATUS","VLANS ALLOWED","NATIVE VLAN"],trunk_data,[14,10,14,14,50,12])

    # ── STATIC ROUTE ──────────────────────────────────────────────────────────
    sr_data = [{"hostname":v(d,"hostname"),"rows":[[r.get("net",NC),r.get("via",NC),r.get("intf",NC),r.get("name",NC)] for r in vl(d,"staticRt")]} for d in devices]
    add_sheet_by_host("STATIC ROUTE","Rotas Estáticas",["Rede / Prefixo","Via (Next-Hop)","Interface","Nome"],sr_data,[22,18,18,24])

    # ── OSPF ──────────────────────────────────────────────────────────────────
    ws_ospf = wb.create_sheet("OSPF")
    nc = 6
    ws_ospf.column_dimensions['A'].width = 2
    for i, w in enumerate([14,16,14,16,40,30]):
        ws_ospf.column_dimensions[get_column_letter(i+2)].width = w
    write_header(ws_ospf, "OSPF", label, nc)
    row = 3
    for d in devices:
        write_hostname_block(ws_ospf, row, v(d,"hostname"), nc)
        row += 1
        write_col_headers(ws_ospf, row, ["Process ID","Router ID","Ref BW","Áreas","Interfaces Ativas","Redistribute"])
        row += 1
        procs = vl(d,"ospfProcs")
        data = [[p.get("pid",NC),p.get("rid",NC),p.get("refBw",NC),", ".join(p.get("areas",[])),", ".join(p.get("activeIfs",[])),", ".join(p.get("redistribute",[]))] for p in procs] if procs else [[NC,NC,NC,NC,NC,NC]]
        row = write_data_rows(ws_ospf, row, data, nc) + 1

    # ── VIZINHANÇA OSPF ───────────────────────────────────────────────────────
    vosp_data = [{"hostname":v(d,"hostname"),"rows":[[r.get("neighborId",NC),r.get("pri",NC),r.get("state",NC),r.get("time",NC),r.get("address",NC),r.get("intf",NC)] for r in vl(d,"ospfNeighbors")]} for d in devices]
    add_sheet_by_host("VIZINHANÇA OSPF","Vizinhança OSPF",["Neighbor ID","Pri","State","Dead/Up Time","Address","Interface"],vosp_data,[16,6,14,12,16,16])

    # ── EIGRP ─────────────────────────────────────────────────────────────────
    eigrp_data = [{"hostname":v(d,"hostname"),"rows":[[r.get("proc",NC),r.get("h",NC),r.get("addr",NC),r.get("intf",NC),r.get("hold",NC),r.get("uptime",NC),r.get("srtt",NC),r.get("rto",NC),r.get("qcnt",NC),r.get("seq",NC)] for r in vl(d,"eigrp")]} for d in devices]
    add_sheet_by_host("EIGRP","EIGRP - Enhanced Interior Gateway Routing Protocol",["Process","H","Address","Interface","Hold","Uptime","SRTT","RTO","Q Cnt","Seq Num"],eigrp_data,[10,6,18,16,8,12,8,8,8,8])

    # ── VIZINHANÇA EIGRP ──────────────────────────────────────────────────────
    veig_data = [{"hostname":v(d,"hostname"),"rows":[[r.get("h",NC),r.get("address",NC),r.get("intf",NC),r.get("hold",NC),r.get("uptime",NC),r.get("srtt",NC),r.get("rto",NC),r.get("qcnt",NC),r.get("seq",NC)] for r in vl(d,"eigrpNeighbors")]} for d in devices]
    add_sheet_by_host("VIZINHANÇA EIGRP","Vizinhança EIGRP",["H","Address","Interface","Hold","Up Time","SRTT","RTO","Q Cnt","Seq"],veig_data,[6,16,16,8,12,8,8,8,8])

    # ── BGP ───────────────────────────────────────────────────────────────────
    bgp_data = [{"hostname":v(d,"hostname"),"rows":[[r.get("rid",NC),r.get("localAs",NC),r.get("neighbor",NC),r.get("v",NC),r.get("as",NC),r.get("msgRcvd",NC),r.get("msgSent",NC),r.get("tblVer",NC),r.get("inQ",NC),r.get("outQ",NC),r.get("upDown",NC),r.get("state",NC)] for r in vl(d,"bgp")]} for d in devices]
    add_sheet_by_host("BGP","BGP - Border Gateway Protocol",["Router ID","Local AS","Neighbor","V","AS","MsgRcvd","MsgSent","TblVer","InQ","OutQ","Up/Down","State/PfxRcd"],bgp_data,[16,10,16,4,10,10,10,10,6,6,12,14])

    # ── VIZINHANÇA BGP ────────────────────────────────────────────────────────
    vbgp_data = [{"hostname":v(d,"hostname"),"rows":[[r.get("rid",NC),r.get("localAs",NC),r.get("neighbor",NC),r.get("v",NC),r.get("as",NC),r.get("msgRcvd",NC),r.get("msgSent",NC),r.get("upDown",NC),r.get("state",NC)] for r in vl(d,"bgpNeighbors")]} for d in devices]
    add_sheet_by_host("VIZINHANÇA BGP","Vizinhança BGP",["Router ID","Local AS","Neighbor","V","AS","MsgRcvd","MsgSent","Up/Down","State/PfxRcd"],vbgp_data,[16,10,16,4,10,10,10,12,14])

    # ── ARP ───────────────────────────────────────────────────────────────────
    arp_data = [{"hostname":v(d,"hostname"),"rows":[[r.get("ip",NC),r.get("mac",NC),r.get("age",NC),r.get("type",NC),r.get("intf",NC)] for r in vl(d,"arpTable")]} for d in devices]
    add_sheet_by_host("ARP","ARP - Address Resolution Protocol",["IP Address","Mac Address","Age(min)","Type","Interface"],arp_data,[18,18,10,10,20])

    # ── MAC ───────────────────────────────────────────────────────────────────
    mac_data = [{"hostname":v(d,"hostname"),"rows":[[r.get("vlan",NC),r.get("mac",NC),r.get("type",NC),r.get("intf",NC)] for r in vl(d,"macTable")]} for d in devices]
    add_sheet_by_host("MAC","MAC - Media Access Control",["VLAN","Mac Address","Type","Interface"],mac_data,[10,18,10,20])

    return wb

# ── Entry point ────────────────────────────────────────────────────────────────
def main():
    data = json.loads(sys.stdin.read())
    mode    = data.get("mode", "matriz")  # "matriz" | "assessment"
    devices = data.get("devices", [])
    label   = data.get("label", "Magneto NTG")

    if mode == "assessment":
        wb = build_assessment_wb(devices, label)
    else:
        # Matriz: uma sheet por device
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for idx, dev in enumerate(devices):
            build_matriz_sheet(wb, dev, idx)

    buf = io.BytesIO()
    wb.save(buf)
    sys.stdout.buffer.write(buf.getvalue())

if __name__ == "__main__":
    main()
